from __future__ import annotations

import re
from io import StringIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO, Callable, NoReturn, Sequence, overload

import polars._reexport as pl
from polars import functions as F
from polars.datatypes import Date, Datetime
from polars.exceptions import NoDataError, ParameterCollisionError
from polars.io.csv.functions import read_csv
from polars.utils.various import normalize_filepath

if TYPE_CHECKING:
    from io import BytesIO
    from typing import Literal

    from polars.type_aliases import SchemaDict


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: str,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: str,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> NoReturn:
    ...


# note: 'ignore' required as mypy thinks that the return value for
# Literal[0] overlaps with the return value for other integers
@overload  # type: ignore[misc]
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: Literal[0] | Sequence[int],
    sheet_name: None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None,
    sheet_name: list[str] | tuple[str],
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    schema_overrides: SchemaDict | None = ...,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int | Sequence[int] | None = None,
    sheet_name: str | list[str] | tuple[str] | None = None,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb"] | None = None,
    xlsx2csv_options: dict[str, Any] | None = None,
    read_csv_options: dict[str, Any] | None = None,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = True,
) -> pl.DataFrame | dict[str, pl.DataFrame]:
    """
    Read Excel (XLSX) spreadsheet data into a DataFrame.

    .. versionadded:: 0.19.4
        Added support for "pyxlsb" engine for reading Excel Binary Workbooks (.xlsb).
    .. versionadded:: 0.19.3
        Added support for "openpyxl" engine, and added `schema_overrides` parameter.

    Parameters
    ----------
    source
        Path to a file or a file-like object (by file-like object, we refer to objects
        that have a `read()` method, such as a file handler (e.g. via builtin `open`
        function) or `BytesIO`).
    sheet_id
        Sheet number(s) to convert (set `0` to load all sheets as DataFrames) and
        return a `{sheetname:frame,}` dict. (Defaults to `1` if neither this nor
        `sheet_name` are specified). Can also take a sequence of sheet numbers.
    sheet_name
        Sheet name(s) to convert; cannot be used in conjunction with `sheet_id`. If more
        than one is given then a `{sheetname:frame,}` dict is returned.
    engine
        Library used to parse the spreadsheet file; defaults to "xlsx2csv" if not set.

        * "xlsx2csv": the fastest engine; converts the data to an in-memory CSV before
          using the native polars `read_csv` method to parse the result. You can
          pass `xlsx2csv_options` and `read_csv_options` to refine the conversion.
        * "openpyxl": this engine is significantly slower than `xlsx2csv` but supports
          additional automatic type inference; potentially useful if you are otherwise
          unable to parse your sheet with the (default) `xlsx2csv` engine in
          conjunction with the `schema_overrides` parameter.
        * "pyxlsb": this engine is used for Excel Binary Workbooks (`.xlsb` files).
          Note that you have to use `schema_overrides` to correctly load date/datetime
          columns (or these will be read as floats representing offset Julian values).

    xlsx2csv_options
        Extra options passed to `xlsx2csv.Xlsx2csv()`,
        e.g. `{"skip_empty_lines": True}`
    read_csv_options
        Extra options passed to :func:`read_csv` for parsing the CSV file returned by
        `xlsx2csv.Xlsx2csv().convert()`
        e.g.: ``{"has_header": False, "new_columns": ["a", "b", "c"],
        "infer_schema_length": None}``
    schema_overrides
        Support type specification or override of one or more columns.
    raise_if_empty
        When there is no data in the sheet,`NoDataError` is raised. If this parameter
        is set to False, an empty DataFrame (with no columns) is returned instead.

    Notes
    -----
    When using the default `xlsx2csv` engine the target Excel sheet is first converted
    to CSV using `xlsx2csv.Xlsx2csv(source).convert()` and then parsed with Polars'
    :func:`read_csv` function. You can pass additional options to `read_csv_options`
    to influence this part of the parsing pipeline.

    Returns
    -------
    DataFrame
        If reading a single sheet.
    dict
        If reading multiple sheets, a "{sheetname: DataFrame, ...}" dict is returned.

    Examples
    --------
    Read the "data" worksheet from an Excel file into a DataFrame.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     sheet_name="data",
    ... )  # doctest: +SKIP

    Read table data from sheet 3 in an Excel workbook as a DataFrame while skipping
    empty lines in the sheet. As sheet 3 does not have a header row and the default
    engine is `xlsx2csv` you can pass the necessary additional settings for this
    to the "read_csv_options" parameter; these will be passed to :func:`read_csv`.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     sheet_id=3,
    ...     xlsx2csv_options={"skip_empty_lines": True},
    ...     read_csv_options={"has_header": False, "new_columns": ["a", "b", "c"]},
    ... )  # doctest: +SKIP

    If the correct datatypes can't be determined you can use `schema_overrides` and/or
    some of the :func:`read_csv` documentation to see which options you can pass to fix
    this issue. For example `"infer_schema_length": None` can be used to read the
    data twice, once to infer the correct output types and once more to then read the
    data with those types. If the types are known in advance then `schema_overrides`
    is the more efficient option.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     read_csv_options={"infer_schema_length": 1000},
    ...     schema_overrides={"dt": pl.Date},
    ... )  # doctest: +SKIP

    The `openpyxl` package can also be used to parse Excel data; it has slightly
    better default type detection, but is slower than `xlsx2csv`. If you have a sheet
    that is better read using this package you can set the engine as "openpyxl" (if you
    use this engine then neither `xlsx2csv_options` nor `read_csv_options` can be set).

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     engine="openpyxl",
    ...     schema_overrides={"dt": pl.Datetime, "value": pl.Int32},
    ... )  # doctest: +SKIP

    """
    if engine and engine != "xlsx2csv":
        if xlsx2csv_options:
            raise ValueError(
                f"cannot specify `xlsx2csv_options` when engine={engine!r}"
            )
        if read_csv_options:
            raise ValueError(
                f"cannot specify `read_csv_options` when engine={engine!r}"
            )

    return _read_spreadsheet(
        sheet_id,
        sheet_name,
        source=source,
        engine=engine,
        engine_options=xlsx2csv_options,
        read_csv_options=read_csv_options,
        schema_overrides=schema_overrides,
        raise_if_empty=raise_if_empty,
    )


@overload
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: str,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: str,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> NoReturn:
    ...


@overload  # type: ignore[misc]
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: Literal[0] | Sequence[int],
    sheet_name: None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


@overload
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None,
    sheet_name: list[str] | tuple[str],
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


def read_ods(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int | Sequence[int] | None = None,
    sheet_name: str | list[str] | tuple[str] | None = None,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = True,
) -> pl.DataFrame | dict[str, pl.DataFrame]:
    """
    Read OpenOffice (ODS) spreadsheet data into a DataFrame.

    Parameters
    ----------
    source
        Path to a file or a file-like object (by file-like object, we refer to objects
        that have a `read()` method, such as a file handler (e.g. via builtin `open`
        function) or `BytesIO`).
    sheet_id
        Sheet number(s) to convert, starting from 1 (set `0` to load *all* worksheets
        as DataFrames) and return a `{sheetname:frame,}` dict. (Defaults to `1` if
        neither this nor `sheet_name` are specified). Can also take a sequence of sheet
        numbers.
    sheet_name
        Sheet name(s) to convert; cannot be used in conjunction with `sheet_id`. If
        more than one is given then a `{sheetname:frame,}` dict is returned.
    schema_overrides
        Support type specification or override of one or more columns.
    raise_if_empty
        When there is no data in the sheet,`NoDataError` is raised. If this parameter
        is set to False, an empty DataFrame (with no columns) is returned instead.

    Returns
    -------
    DataFrame, or a `{sheetname: DataFrame, ...}` dict if reading multiple sheets.

    Examples
    --------
    Read the "data" worksheet from an OpenOffice spreadsheet file into a DataFrame.

    >>> pl.read_ods(
    ...     source="test.ods",
    ...     sheet_name="data",
    ... )  # doctest: +SKIP

    If the correct dtypes can't be determined, use the `schema_overrides` parameter
    to specify them.

    >>> pl.read_ods(
    ...     source="test.ods",
    ...     sheet_id=3,
    ...     schema_overrides={"dt": pl.Date},
    ...     raise_if_empty=False,
    ... )  # doctest: +SKIP

    """
    return _read_spreadsheet(
        sheet_id,
        sheet_name,
        source=source,
        engine="ods",
        engine_options={},
        read_csv_options={},
        schema_overrides=schema_overrides,
        raise_if_empty=raise_if_empty,
    )


def _read_spreadsheet(
    sheet_id: int | Sequence[int] | None,
    sheet_name: str | list[str] | tuple[str] | None,
    source: str | BytesIO | Path | BinaryIO | bytes,
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb", "ods"] | None,
    engine_options: dict[str, Any] | None = None,
    read_csv_options: dict[str, Any] | None = None,
    schema_overrides: SchemaDict | None = None,
    *,
    raise_if_empty: bool = True,
) -> pl.DataFrame | dict[str, pl.DataFrame]:
    if isinstance(source, (str, Path)):
        source = normalize_filepath(source)

    if engine is None:
        if (src := str(source).lower()).endswith(".ods"):
            engine = "ods"
        else:
            engine = "pyxlsb" if src.endswith(".xlsb") else "xlsx2csv"

    # establish the reading function, parser, and available worksheets
    reader_fn, parser, worksheets = _initialise_spreadsheet_parser(
        engine, source, engine_options or {}
    )
    try:
        # parse data from the indicated sheet(s)
        sheet_names, return_multi = _get_sheet_names(sheet_id, sheet_name, worksheets)
        parsed_sheets = {
            name: reader_fn(
                parser=parser,
                sheet_name=name,
                read_csv_options=read_csv_options,
                schema_overrides=schema_overrides,
                raise_if_empty=raise_if_empty,
            )
            for name in sheet_names
        }
    finally:
        if hasattr(parser, "close"):
            parser.close()

    if not parsed_sheets:
        param, value = ("id", sheet_id) if sheet_name is None else ("name", sheet_name)
        raise ValueError(f"no matching sheets found when `sheet_{param}` is {value!r}")

    if return_multi:
        return parsed_sheets
    return next(iter(parsed_sheets.values()))


def _get_sheet_names(
    sheet_id: int | Sequence[int] | None,
    sheet_name: str | list[str] | tuple[str] | None,
    worksheets: list[dict[str, Any]],
) -> tuple[list[str], bool]:
    """Establish sheets to read; indicate if we are returning a dict frames."""
    if sheet_id is not None and sheet_name is not None:
        raise ValueError(
            f"cannot specify both `sheet_name` ({sheet_name!r}) and `sheet_id` ({sheet_id!r})"
        )
    sheet_names = []
    if sheet_id is None and sheet_name is None:
        sheet_names.append(worksheets[0]["name"])
        return_multi = False
    elif sheet_id == 0:
        sheet_names.extend(ws["name"] for ws in worksheets)
        return_multi = True
    else:
        return_multi = (
            (isinstance(sheet_name, Sequence) and not isinstance(sheet_name, str))
            or isinstance(sheet_id, Sequence)
            or sheet_id == 0
        )
        if names := (
            (sheet_name,) if isinstance(sheet_name, str) else sheet_name or ()
        ):
            known_sheet_names = {ws["name"] for ws in worksheets}
            for name in names:
                if name not in known_sheet_names:
                    raise ValueError(
                        f"no matching sheet found when `sheet_name` is {name!r}"
                    )
                sheet_names.append(name)
        else:
            ids = (sheet_id,) if isinstance(sheet_id, int) else sheet_id or ()
            sheet_names_by_idx = {
                idx: ws["name"]
                for idx, ws in enumerate(worksheets, start=1)
                if (sheet_id == 0 or ws["index"] in ids or ws["name"] in names)
            }
            for idx in ids:
                if (name := sheet_names_by_idx.get(idx)) is None:  # type: ignore[assignment]
                    raise ValueError(
                        f"no matching sheet found when `sheet_id` is {idx}"
                    )
                sheet_names.append(name)
    return sheet_names, return_multi


def _initialise_spreadsheet_parser(
    engine: Literal["xlsx2csv", "openpyxl", "pyxlsb", "ods"],
    source: str | BytesIO | Path | BinaryIO | bytes,
    engine_options: dict[str, Any],
) -> tuple[Callable[..., pl.DataFrame], Any, list[dict[str, Any]]]:
    """Instantiate the indicated spreadsheet parser and establish related properties."""
    if engine == "xlsx2csv":  # default
        try:
            import xlsx2csv
        except ImportError:
            raise ModuleNotFoundError(
                "required package not installed" "\n\nPlease run: pip install xlsx2csv"
            ) from None

        # establish sensible defaults for unset options
        for option, value in {
            "exclude_hidden_sheets": False,
            "skip_empty_lines": False,
            "skip_hidden_rows": False,
            "floatformat": "%f",
        }.items():
            engine_options.setdefault(option, value)

        parser = xlsx2csv.Xlsx2csv(source, **engine_options)
        sheets = parser.workbook.sheets
        return _read_spreadsheet_xlsx2csv, parser, sheets

    elif engine == "openpyxl":
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "required package not installed" "\n\nPlease run: pip install openpyxl"
            ) from None
        parser = openpyxl.load_workbook(source, data_only=True, **engine_options)
        sheets = [{"index": i + 1, "name": ws.title} for i, ws in enumerate(parser)]
        return _read_spreadsheet_openpyxl, parser, sheets

    elif engine == "pyxlsb":
        try:
            import pyxlsb
        except ImportError:
            raise ImportError(
                "required package not installed" "\n\nPlease run: pip install pyxlsb"
            ) from None
        try:
            parser = pyxlsb.open_workbook(source, **engine_options)
        except KeyError as err:
            if "no item named 'xl/_rels/workbook.bin.rels'" in str(err):
                raise TypeError(f"invalid Excel Binary Workbook: {source!r}") from None
            raise
        sheets = [
            {"index": i + 1, "name": name} for i, name in enumerate(parser.sheets)
        ]
        return _read_spreadsheet_pyxlsb, parser, sheets

    elif engine == "ods":
        try:
            import ezodf
        except ImportError:
            raise ImportError(
                "required package not installed"
                "\n\nPlease run: pip install ezodf lxml"
            ) from None
        parser = ezodf.opendoc(source, **engine_options)
        sheets = [
            {"index": i + 1, "name": ws.name} for i, ws in enumerate(parser.sheets)
        ]
        return _read_spreadsheet_ods, parser, sheets

    raise NotImplementedError(f"unrecognized engine: {engine!r}")


def _csv_buffer_to_frame(
    csv: StringIO,
    separator: str,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Translate StringIO buffer containing delimited data as a DataFrame."""
    # handle (completely) empty sheet data
    if csv.tell() == 0:
        if raise_if_empty:
            raise NoDataError(
                "empty Excel sheet"
                "\n\nIf you want to read this as an empty DataFrame, set `raise_if_empty=False`."
            )
        return pl.DataFrame()

    if read_csv_options is None:
        read_csv_options = {}
    if schema_overrides:
        if (csv_dtypes := read_csv_options.get("dtypes", {})) and set(
            csv_dtypes
        ).intersection(schema_overrides):
            raise ParameterCollisionError(
                "cannot specify columns in both `schema_overrides` and `read_csv_options['dtypes']`"
            )
        read_csv_options = read_csv_options.copy()
        read_csv_options["dtypes"] = {**csv_dtypes, **schema_overrides}

    # otherwise rewind the buffer and parse as csv
    csv.seek(0)
    df = read_csv(
        csv,
        separator=separator,
        **read_csv_options,
    )
    return _drop_unnamed_null_columns(df)


def _drop_unnamed_null_columns(df: pl.DataFrame) -> pl.DataFrame:
    """If DataFrame contains unnamed columns that contain only nulls, drop them."""
    null_cols = []
    for col_name in df.columns:
        # note that if multiple unnamed columns are found then all but
        # the first one will be ones will be named as "_duplicated_{n}"
        if col_name == "" or re.match(r"_duplicated_\d+$", col_name):
            if df[col_name].null_count() == len(df):
                null_cols.append(col_name)
    if null_cols:
        df = df.drop(*null_cols)
    return df


def _read_spreadsheet_ods(
    parser: Any,
    sheet_name: str | None,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'ezodf' library to read data from the given worksheet."""
    sheets = parser.sheets
    if sheet_name is not None:
        ws = next((s for s in sheets if s.name == sheet_name), None)
        if ws is None:
            raise ValueError(f"sheet {sheet_name!r} not found")
    else:
        ws = sheets[0]

    row_data = []
    found_row_data = False
    for row in ws.rows():
        row_values = [c.value for c in row]
        if found_row_data or (found_row_data := any(v is not None for v in row_values)):
            row_data.append(row_values)

    overrides = {}
    strptime_cols = {}
    headers: list[str] = []

    if not row_data:
        df = pl.DataFrame()
    else:
        for idx, name in enumerate(row_data[0]):
            headers.append(name or (f"_duplicated_{idx}" if headers else ""))

        trailing_null_row = all(v is None for v in row_data[-1])
        row_data = row_data[1 : -1 if trailing_null_row else None]

        if schema_overrides:
            for nm, dtype in schema_overrides.items():
                if dtype in (Datetime, Date):
                    strptime_cols[nm] = dtype
                else:
                    overrides[nm] = dtype

        df = pl.DataFrame(
            row_data,
            orient="row",
            schema=headers,
            schema_overrides=overrides,
        )

    if raise_if_empty and len(df) == 0 and len(df.columns) == 0:
        raise NoDataError(
            "empty Excel sheet"
            "\n\nIf you want to read this as an empty DataFrame, set `raise_if_empty=False`."
        )

    if strptime_cols:
        df = df.with_columns(
            (
                F.col(nm).str.replace("[T ]00:00:00$", "")
                if dtype == Date
                else F.col(nm)
            ).str.strptime(
                dtype  # type: ignore[arg-type]
            )
            for nm, dtype in strptime_cols.items()
        )
    df.columns = headers
    return _drop_unnamed_null_columns(df)


def _read_spreadsheet_openpyxl(
    parser: Any,
    sheet_name: str | None,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'openpyxl' library to read data from the given worksheet."""
    ws = parser[sheet_name]

    # prefer detection of actual table objects; otherwise read
    # data in the used worksheet range, dropping null columns
    header: list[str | None] = []
    if tables := getattr(ws, "tables", None):
        table = next(iter(tables.values()))
        rows = list(ws[table.ref])
        header.extend(cell.value for cell in rows.pop(0))
        if table.totalsRowCount:
            rows = rows[: -table.totalsRowCount]
        rows_iter = iter(rows)
    else:
        rows_iter = ws.iter_rows()
        for row in rows_iter:
            row_values = [cell.value for cell in row]
            if any(v is not None for v in row_values):
                header.extend(row_values)
                break

    series_data = [
        pl.Series(name, [cell.value for cell in column_data])
        for name, column_data in zip(header, zip(*rows_iter))
        if name
    ]
    df = pl.DataFrame(
        {s.name: s for s in series_data},
        schema_overrides=schema_overrides,
    )
    if raise_if_empty and len(df) == 0 and len(df.columns) == 0:
        raise NoDataError(
            "empty Excel sheet"
            "\n\nIf you want to read this as an empty DataFrame, set `raise_if_empty=False`."
        )
    return _drop_unnamed_null_columns(df)


def _read_spreadsheet_pyxlsb(
    parser: Any,
    sheet_name: str | None,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    from pyxlsb import convert_date

    ws = parser.get_sheet(sheet_name)
    try:
        # establish header/data rows
        header: list[str | None] = []
        rows_iter = ws.rows()
        for row in rows_iter:
            row_values = [cell.v for cell in row]
            if any(v is not None for v in row_values):
                header.extend(row_values)
                break

        # load data rows as series
        series_data = [
            pl.Series(name, [cell.v for cell in column_data])
            for name, column_data in zip(header, zip(*rows_iter))
            if name
        ]
    finally:
        ws.close()

    if schema_overrides:
        for idx, s in enumerate(series_data):
            if schema_overrides.get(s.name) in (Datetime, Date):
                series_data[idx] = s.map_elements(convert_date)

    df = pl.DataFrame(
        {s.name: s for s in series_data},
        schema_overrides=schema_overrides,
    )
    if raise_if_empty and len(df) == 0 and len(df.columns) == 0:
        raise NoDataError(
            "empty Excel sheet"
            "\n\nIf you want to read this as an empty DataFrame, set `raise_if_empty=False`."
        )
    return _drop_unnamed_null_columns(df)


def _read_spreadsheet_xlsx2csv(
    parser: Any,
    sheet_name: str | None,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'xlsx2csv' library to read data from the given worksheet."""
    csv_buffer = StringIO()
    parser.convert(
        outfile=csv_buffer,
        sheetname=sheet_name,
    )
    if read_csv_options is None:
        read_csv_options = {}
    read_csv_options.setdefault("truncate_ragged_lines", True)

    return _csv_buffer_to_frame(
        csv_buffer,
        separator=",",
        read_csv_options=read_csv_options,
        schema_overrides=schema_overrides,
        raise_if_empty=raise_if_empty,
    )
