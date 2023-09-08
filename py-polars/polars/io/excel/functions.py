from __future__ import annotations

import re
from io import StringIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO, NoReturn, Sequence, overload

import polars._reexport as pl
from polars import functions as F
from polars.datatypes import Date, Datetime
from polars.exceptions import NoDataError
from polars.io.csv.functions import read_csv
from polars.utils.various import normalize_filepath

if TYPE_CHECKING:
    from io import BytesIO
    from typing import Literal

    from polars.type_aliases import SchemaDict


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: str,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None = ...,
    sheet_name: None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: str,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> NoReturn:
    ...


# note: 'ignore' required as mypy thinks that the return value for
# Literal[0] overlaps with the return value for other integers
@overload  # type: ignore[misc]
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: Literal[0] | Sequence[int],
    sheet_name: None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: None,
    sheet_name: list[str] | tuple[str],
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> dict[str, pl.DataFrame]:
    ...


@overload
def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int,
    sheet_name: None = ...,
    xlsx2csv_options: dict[str, Any] | None = ...,
    read_csv_options: dict[str, Any] | None = ...,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = ...,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = ...,
) -> pl.DataFrame:
    ...


def read_excel(
    source: str | BytesIO | Path | BinaryIO | bytes,
    *,
    sheet_id: int | Sequence[int] | None = None,
    sheet_name: str | list[str] | tuple[str] | None = None,
    xlsx2csv_options: dict[str, Any] | None = None,
    read_csv_options: dict[str, Any] | None = None,
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None = None,
    schema_overrides: SchemaDict | None = None,
    raise_if_empty: bool = True,
) -> pl.DataFrame | dict[str, pl.DataFrame]:
    """
    Read Excel (XLSX) sheet into a DataFrame.

    If using the ``xlsx2csv`` engine, converts an Excel sheet with
    ``xlsx2csv.Xlsx2csv().convert()`` to CSV and parses the CSV output with
    :func:`read_csv`. You can pass additional options to ``read_csv_options`` to
    influence parsing behaviour.

    When using the ``openpyxl`` engine, reads an Excel sheet with
    ``openpyxl.load_workbook(source)``.

    Parameters
    ----------
    source
        Path to a file or a file-like object (by file-like object, we refer to objects
        that have a ``read()`` method, such as a file handler (e.g. via builtin ``open``
        function) or ``BytesIO``).
    sheet_id
        Sheet number to convert (set ``0`` to load all sheets as DataFrames) and return
        a ``{sheetname:frame,}`` dict. (Defaults to `1` if neither this nor `sheet_name`
        are specified). Can also take a sequence of sheet numbers.
    sheet_name
        Sheet name()s to convert; cannot be used in conjunction with `sheet_id`. If more
        than one is given then a ``{sheetname:frame,}`` dict is returned.
    xlsx2csv_options
        Extra options passed to ``xlsx2csv.Xlsx2csv()``,
        e.g. ``{"skip_empty_lines": True}``
    read_csv_options
        Extra options passed to :func:`read_csv` for parsing the CSV file returned by
        ``xlsx2csv.Xlsx2csv().convert()``
        e.g.: ``{"has_header": False, "new_columns": ["a", "b", "c"],
        "infer_schema_length": None}``
    engine
        Library used to parse Excel, either openpyxl or xlsx2csv (default is xlsx2csv).
        Please note that xlsx2csv converts first to csv, making type inference worse
        than openpyxl. To remedy that, you can use the extra options defined on
        `xlsx2csv_options` and `read_csv_options`
    schema_overrides
        Support type specification or override of one or more columns.
    raise_if_empty
        When there is no data in the sheet,``NoDataError`` is raised. If this parameter
        is set to False, an empty DataFrame (with no columns) is returned instead.

    Returns
    -------
    DataFrame, or a dictionary of sheetname to DataFrame if reading multiple sheets.

    Examples
    --------
    Read the "data" worksheet from an Excel file into a DataFrame.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     sheet_name="data",
    ... )  # doctest: +SKIP

    Read sheet 3 from Excel sheet file to a DataFrame while skipping empty lines in the
    sheet. As sheet 3 does not have header row, pass the needed settings to
    :func:`read_csv`.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     sheet_id=3,
    ...     xlsx2csv_options={"skip_empty_lines": True},
    ...     read_csv_options={"has_header": False, "new_columns": ["a", "b", "c"]},
    ... )  # doctest: +SKIP

    If the correct datatypes can't be determined by polars, look at the :func:`read_csv`
    documentation to see which options you can pass to fix this issue. For example
    ``"infer_schema_length": None`` can be used to read the data twice, once to infer
    the correct output types and once to  convert the input to the correct types.
    When `"infer_schema_length": 1000``, only the first 1000 lines are read twice.

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     read_csv_options={"infer_schema_length": None},
    ...     schema_overrides={"dt": pl.Date},
    ... )  # doctest: +SKIP

    The ``openpyxl`` engine can also be used to provide automatic type inference.
    To do so, specify the right engine (`xlsx2csv_options` and `read_csv_options`
    will be ignored):

    >>> pl.read_excel(
    ...     source="test.xlsx",
    ...     engine="openpyxl",
    ...     schema_overrides={"dt": pl.Datetime, "value": pl.Int32},
    ... )  # doctest: +SKIP

    """
    if sheet_id is not None and sheet_name is not None:
        raise ValueError(
            f"cannot specify both `sheet_name` ({sheet_name!r}) and `sheet_id` ({sheet_id!r})"
        )

    if xlsx2csv_options is None:
        xlsx2csv_options = {}

    if read_csv_options is None:
        read_csv_options = {"truncate_ragged_lines": True}
    elif "truncate_ragged_lines" not in read_csv_options:
        read_csv_options["truncate_ragged_lines"] = True

    # establish the reading function, parser, and available worksheets
    reader_fn, parser, worksheets = _initialise_excel_parser(
        engine, source, xlsx2csv_options
    )

    # use the parser to read data from one or more sheets
    if (
        sheet_id == 0
        or isinstance(sheet_id, Sequence)
        or (sheet_name and not isinstance(sheet_name, str))
    ):
        # read multiple sheets by id
        sheet_ids = sheet_id or ()
        sheet_names = sheet_name or ()
        return {
            sheet["name"]: reader_fn(
                parser=parser,
                sheet_id=sheet["index"],
                sheet_name=None,
                read_csv_options=read_csv_options,
                schema_overrides=schema_overrides,
                raise_if_empty=raise_if_empty,
            )
            for sheet in worksheets
            if sheet_id == 0 or sheet["index"] in sheet_ids or sheet["name"] in sheet_names  # type: ignore[operator]
        }
    else:
        # read a specific sheet by id or name
        if sheet_name is None:
            sheet_id = sheet_id or 1
        return reader_fn(
            parser=parser,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
            read_csv_options=read_csv_options,
            schema_overrides=schema_overrides,
            raise_if_empty=raise_if_empty,
        )


def _initialise_excel_parser(
    engine: Literal["xlsx2csv", "openpyxl", "odf"] | None,
    source: str | BytesIO | Path | BinaryIO | bytes,
    xlsx2csv_options: dict[str, Any],
) -> tuple[Any, Any, list[dict[str, Any]]]:
    """Instantiate the indicated Excel parser and establish related properties."""
    if isinstance(source, (str, Path)):
        source = normalize_filepath(source)
    if engine is None and str(source).lower().endswith(".ods"):
        engine = "odf"

    if engine == "xlsx2csv" or engine is None:  # default
        try:
            import xlsx2csv
        except ImportError:
            raise ModuleNotFoundError(
                "Required package not installed\n\nPlease run: `pip install xlsx2csv`"
            ) from None
        parser = xlsx2csv.Xlsx2csv(source, **xlsx2csv_options)
        sheets = parser.workbook.sheets
        return _read_excel_sheet_xlsx2csv, parser, sheets

    elif engine == "openpyxl":
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Required package not installed\n\nPlease run `pip install openpyxl`"
            ) from None
        parser = openpyxl.load_workbook(source, data_only=True)
        sheets = [{"index": i + 1, "name": ws.title} for i, ws in enumerate(parser)]
        return _read_excel_sheet_openpyxl, parser, sheets

    elif engine == "odf":
        try:
            import ezodf
        except ImportError:
            raise ImportError(
                "Required package not installed\n\nPlease run `pip install ezodf lxml`"
            ) from None
        parser = ezodf.opendoc(source)
        sheets = [
            {"index": i + 1, "name": ws.name} for i, ws in enumerate(parser.sheets)
        ]
        return _read_excel_sheet_odf, parser, sheets

    raise NotImplementedError(f"Unrecognised engine: {engine!r}")


def _csv_buffer_to_frame(
    csv: StringIO,
    separator: str,
    read_csv_options: dict[str, Any],
    schema_overrides: SchemaDict | None,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Translate StringIO buffer containing delimited data as a DataFrame."""
    # handle (completely) empty sheet data
    if csv.tell() == 0:
        if raise_if_empty:
            raise NoDataError(
                "Empty Excel sheet; if you want to read this as "
                "an empty DataFrame, set `raise_if_empty=False`"
            )
        return pl.DataFrame()

    # otherwise rewind the buffer and parse as csv
    csv.seek(0)
    df = read_csv(csv, separator=separator, dtypes=schema_overrides, **read_csv_options)
    return _drop_unnamed_null_columns(df)


def _drop_unnamed_null_columns(df: pl.DataFrame) -> pl.DataFrame:
    """If DataFrame contains unnamed columns that contain only nulls, drop them."""
    null_cols = []
    for col_name in df.columns:
        # note that if multiple unnamed columns are found then all but
        # the first one will be ones will be named as "_duplicated_{n}"
        if col_name == "" or re.match(r"_duplicated_\d+$", col_name):
            if df[col_name].null_count() == len(df):
                null_cols.append(col_name)
    if null_cols:
        df = df.drop(*null_cols)
    return df


def _read_excel_sheet_odf(
    parser: Any,
    sheet_id: int | None,
    sheet_name: str | None,
    read_csv_options: dict[str, Any],
    schema_overrides: SchemaDict | None,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'ezodf' library to read data from the given worksheet."""
    sheets = parser.sheets
    if sheet_id is not None:
        ws = sheets[sheet_id - 1]
    elif sheet_name is not None:
        ws = next((s for s in sheets if s.name == sheet_name), None)
        if ws is None:
            raise ValueError(f"Sheet {sheet_name!r} not found")
    else:
        ws = sheets[0]

    row_data = []
    found_row_data = False
    for row in ws.rows():
        row_values = [c.value for c in row]
        if found_row_data or (found_row_data := any(v is not None for v in row_values)):
            row_data.append(row_values)

    headers: list[str] = []
    for idx, name in enumerate(row_data[0]):
        headers.append(name or (f"_duplicated_{idx}" if headers else ""))

    trailing_null_row = all(v is None for v in row_data[-1])
    row_data = row_data[1 : -1 if trailing_null_row else None]

    overrides = {}
    strptime_cols = {}
    if schema_overrides:
        for nm, dtype in schema_overrides.items():
            if dtype in (Datetime, Date):
                strptime_cols[nm] = dtype
            else:
                overrides[nm] = dtype

    df = pl.DataFrame(
        row_data,
        orient="row",
        schema=headers,
        schema_overrides=overrides,
    )
    if raise_if_empty and len(df) == 0 and len(df.columns) == 0:
        raise NoDataError(
            "Empty Excel sheet; if you want to read this as "
            "an empty DataFrame, set `raise_if_empty=False`"
        )

    if strptime_cols:
        df = df.with_columns(
            F.col(nm).str.strptime(dtype)  # type: ignore[arg-type]
            for nm, dtype in strptime_cols.items()
        )

    df.columns = headers
    return _drop_unnamed_null_columns(df)


def _read_excel_sheet_openpyxl(
    parser: Any,
    sheet_id: int | None,
    sheet_name: str | None,
    read_csv_options: dict[str, Any] | None,
    schema_overrides: SchemaDict | None,
    *,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'openpyxl' library to read data from the given worksheet."""
    # read requested sheet if provided on kwargs, otherwise read active sheet
    if sheet_name is not None:
        ws = parser[sheet_name]
    elif sheet_id is not None:
        ws = parser.worksheets[sheet_id - 1]
    else:
        ws = parser.active

    # prefer detection of actual table objects; otherwise read
    # data in the used worksheet range, dropping null columns
    header: list[str | None] = []
    if tables := getattr(ws, "tables", None):
        table = next(iter(tables.values()))
        rows = list(ws[table.ref])
        header.extend(cell.value for cell in rows.pop(0))
        if table.totalsRowCount:
            rows = rows[: -table.totalsRowCount]
        rows_iter = iter(rows)
    else:
        rows_iter = ws.iter_rows()
        for row in rows_iter:
            row_values = [cell.value for cell in row]
            if any(v is not None for v in row_values):
                header.extend(row_values)
                break

    series_data = [
        pl.Series(name, [cell.value for cell in column_data])
        for name, column_data in zip(header, zip(*rows_iter))
    ]
    df = pl.DataFrame(
        {s.name: s for s in series_data if s.name},
        schema_overrides=schema_overrides,
    )
    if raise_if_empty and len(df) == 0 and len(df.columns) == 0:
        raise NoDataError(
            "Empty Excel sheet; if you want to read this as "
            "an empty DataFrame, set `raise_if_empty=False`"
        )
    return _drop_unnamed_null_columns(df)


def _read_excel_sheet_xlsx2csv(
    parser: Any,
    sheet_id: int | None,
    sheet_name: str | None,
    read_csv_options: dict[str, Any],
    *,
    schema_overrides: SchemaDict | None,
    raise_if_empty: bool,
) -> pl.DataFrame:
    """Use the 'xlsx2csv' library to read data from the given worksheet."""
    csv_buffer = StringIO()
    parser.convert(
        outfile=csv_buffer,
        sheetid=sheet_id,
        sheetname=sheet_name,
    )
    return _csv_buffer_to_frame(
        csv_buffer,
        separator=",",
        read_csv_options=read_csv_options,
        schema_overrides=schema_overrides,
        raise_if_empty=raise_if_empty,
    )
